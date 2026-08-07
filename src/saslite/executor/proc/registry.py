"""PROC handler registry and built-in PROC implementations."""

from __future__ import annotations

import io
from typing import Any

import pandas as pd

from saslite.ast.proc import ProcNode, VarListNode, ByNode, ClassNode, FreqTableSpec
from saslite.ast.data_step import DatasetRefNode
from saslite.executor.expression_eval import ExpressionEvaluator
from saslite.functions import build_default_registry
from saslite.runtime.dataset import Dataset
from saslite.runtime.execution_result import StepResult
from saslite.runtime.formatting import csv_dataframe
from saslite.runtime.types import sas_bool
from saslite.session.session import Session
from saslite.diagnostics.reporter import Reporter


def handle_proc_print(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC PRINT — display dataset contents."""
    data_name = proc.options.get("DATA", "")
    if not data_name:
        return StepResult(success=False, error="PROC PRINT requires DATA=")

    try:
        if "." in data_name:
            parts = data_name.split(".", 1)
            ds = session.get_dataset(parts[0], parts[1])
        else:
            ds = session.get_dataset("WORK", data_name)
    except KeyError:
        return StepResult(success=False, error=f"Dataset {data_name} not found")

    # Collect statements
    var_stmt = None
    by_stmt = None
    sum_vars = []
    id_vars = []

    for stmt in proc.statements:
        if isinstance(stmt, VarListNode):
            if var_stmt is None:
                var_stmt = stmt
            # SUM variables
            if hasattr(proc.statements, '__iter__'):
                pass  # handled below
        elif isinstance(stmt, ByNode):
            by_stmt = stmt

    # Separate VAR and SUM statements by looking at proc.statements order
    # VAR statements come from VarListNode, SUM also comes from VarListNode
    # We distinguish by position: first VarListNode = VAR, subsequent = SUM
    var_nodes = [s for s in proc.statements if isinstance(s, VarListNode)]
    if var_nodes:
        var_stmt = var_nodes[0]
    if len(var_nodes) > 1:
        sum_vars = [v.upper() for v in var_nodes[1].variables]

    # Check for ID statement (also a ByNode in current impl)
    by_nodes = [s for s in proc.statements if isinstance(s, ByNode)]
    by_vars = [v.upper() for v in by_nodes[0].variables] if by_nodes else []

    if var_stmt and var_stmt.variables:
        # Always include BY columns even if not in VAR
        want_cols = set(v.upper() for v in var_stmt.variables)
        want_cols.update(by_vars)
        cols = [c for c in ds.data.columns if c.upper() in want_cols]
        display_df = ds.data[cols].copy()
    else:
        display_df = ds.data.copy()

    buf = io.StringIO()
    buf.write(f"\n{'=' * 60}\n")
    buf.write(f"  DATA: {ds.metadata.libref}.{ds.metadata.member_name}\n")
    buf.write(f"  Observations: {ds.nrow}  Variables: {ds.ncol}\n")
    buf.write(f"{'=' * 60}\n\n")

    if by_vars:
        # Grouped display
        # Find actual column names (case-insensitive)
        col_map = {c.upper(): c for c in display_df.columns}
        actual_by = [col_map[v] for v in by_vars if v in col_map]
        if actual_by:
            display_df = display_df.sort_values(by=actual_by).reset_index(drop=True)
            obs_num = 1
            grand_totals: dict[str, float] = {}
            for group_key, group_df in display_df.groupby(actual_by, sort=False):
                # Group header
                if isinstance(group_key, tuple):
                    header = "  ".join(f"{a}={k}" for a, k in zip(actual_by, group_key))
                else:
                    header = f"{actual_by[0]}={group_key}"
                buf.write(f"\n  --- {header} ---\n")

                # Print group rows
                group_display = group_df.drop(columns=actual_by, errors="ignore").copy()
                group_display.index = range(obs_num, obs_num + len(group_display))
                group_display.index.name = "Obs"
                buf.write(group_display.to_string())
                buf.write("\n")
                obs_num += len(group_display)

                # Accumulate SUM vars
                for sv in sum_vars:
                    actual_sv = col_map.get(sv)
                    if actual_sv and actual_sv in group_df.columns:
                        try:
                            total = group_df[actual_sv].sum()
                            buf.write(f"\n  Sum: {actual_sv} = {total}\n")
                            grand_totals[actual_sv] = grand_totals.get(actual_sv, 0) + total
                        except (TypeError, ValueError):
                            pass

            # Grand total
            if grand_totals:
                buf.write(f"\n  Grand Total:\n")
                for sv, total in grand_totals.items():
                    buf.write(f"    {sv} = {total}\n")
        else:
            # BY vars not found in data
            _print_plain(buf, display_df, sum_vars)
    else:
        _print_plain(buf, display_df, sum_vars)

    output = buf.getvalue()
    reporter.log(output)

    return StepResult(
        success=True,
        dataset_name=data_name,
        rows_affected=ds.nrow,
        output_messages=[output],
    )


def _print_plain(buf: io.StringIO, display_df: pd.DataFrame, sum_vars: list[str]) -> None:
    """Print dataframe with optional SUM totals."""
    display = display_df.copy()
    display.index = range(1, len(display) + 1)
    display.index.name = "Obs"
    buf.write(display.to_string())
    buf.write("\n")

    # SUM totals
    col_map = {c.upper(): c for c in display_df.columns}
    for sv in sum_vars:
        actual = col_map.get(sv)
        if actual and actual in display_df.columns:
            try:
                total = display_df[actual].sum()
                buf.write(f"\n  Sum: {actual} = {total}\n")
            except (TypeError, ValueError):
                pass


def handle_proc_sort(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC SORT — sort dataset."""
    from saslite.ast.data_step import DatasetRefNode

    data_ref = proc.options.get("DATA", "")
    out_ref = proc.options.get("OUT", data_ref)

    if not data_ref:
        return StepResult(success=False, error="PROC SORT requires DATA=")

    # Handle DatasetRefNode or string
    if isinstance(data_ref, DatasetRefNode):
        data_libref = data_ref.libref
        data_name = data_ref.name
        data_options = data_ref.options
    elif isinstance(data_ref, str):
        if "." in data_ref:
            parts = data_ref.split(".", 1)
            data_libref = parts[0]
            data_name = parts[1]
        else:
            data_libref = "WORK"
            data_name = data_ref
        data_options = {}
    else:
        return StepResult(success=False, error="Invalid DATA= specification")

    try:
        ds = session.get_dataset(data_libref, data_name)
    except KeyError:
        return StepResult(success=False, error=f"Dataset {data_libref}.{data_name} not found")

    # Apply dataset options (KEEP, DROP, WHERE, RENAME).  Keep a Dataset in
    # parallel with the frame so PROC SORT does not re-infer and lose SAS
    # attributes (especially for empty data sets and all-missing columns).
    working_ds = ds.copy()
    df = working_ds.data

    if "KEEP" in data_options:
        keep_vars = [v.upper() for v in data_options["KEEP"]]
        col_map = {c.upper(): c for c in df.columns}
        actual_keep = [col_map[v] for v in keep_vars if v in col_map]
        working_ds = working_ds.select_columns(actual_keep)
        df = working_ds.data

    if "DROP" in data_options:
        drop_vars = [v.upper() for v in data_options["DROP"]]
        col_map = {c.upper(): c for c in df.columns}
        actual_drop = [col_map[v] for v in drop_vars if v in col_map]
        working_ds = working_ds.select_columns([
            column for column in df.columns if column not in actual_drop
        ])
        df = working_ds.data

    if "RENAME" in data_options:
        rename_map = {}
        col_map = {c.upper(): c for c in df.columns}
        for old_name, new_name in data_options["RENAME"].items():
            old_upper = old_name.upper()
            if old_upper in col_map:
                rename_map[col_map[old_upper]] = new_name.upper()
        working_ds = working_ds.rename_columns(rename_map)
        df = working_ds.data

    # Get BY variables
    by_vars = []
    descending = []
    for stmt in proc.statements:
        if isinstance(stmt, ByNode):
            by_vars = stmt.variables
            break

    if not by_vars:
        return StepResult(success=False, error="PROC SORT requires BY statement")

    # Resolve BY vars to actual column names (case-insensitive)
    col_map = {c.upper(): c for c in df.columns}
    resolved_by = []
    for bv in by_vars:
        actual = col_map.get(bv.upper())
        if actual:
            resolved_by.append(actual)
        else:
            return StepResult(success=False, error=f"Variable {bv} not found in dataset")

    # Sort
    asc = proc.options.get("_ascending", [True] * len(resolved_by))
    tie_cols = [c for c in df.columns if c not in resolved_by]
    sorted_df = df.copy()
    sorted_df["__sas_sort_ord__"] = range(len(sorted_df))
    sorted_df = sorted_df.sort_values(
        by=resolved_by + tie_cols + ["__sas_sort_ord__"],
        ascending=asc + ([True] * len(tie_cols)) + [True],
        kind="mergesort",
    ).reset_index(drop=True)

    # Apply NODUPKEY or NODUPRECS
    nodupkey = proc.options.get("NODUPKEY", False)
    noduprecs = proc.options.get("NODUPRECS", False)

    if nodupkey:
        sorted_df = sorted_df.drop_duplicates(subset=resolved_by, keep="first").reset_index(drop=True)
    elif noduprecs:
        sorted_df = sorted_df.drop_duplicates(keep="first").reset_index(drop=True)

    if "__sas_sort_ord__" in sorted_df.columns:
        sorted_df = sorted_df.drop(columns="__sas_sort_ord__")

    # Resolve output libref and name
    if isinstance(out_ref, DatasetRefNode):
        out_libref = out_ref.libref
        out_member = out_ref.name
    elif isinstance(out_ref, str):
        if "." in out_ref:
            parts = out_ref.split(".", 1)
            out_libref = parts[0]
            out_member = parts[1]
        else:
            out_libref = "WORK"
            out_member = out_ref
    else:
        out_libref = data_libref
        out_member = data_name

    output_metadata = working_ds.metadata.copy()
    output_metadata.libref = out_libref
    output_metadata.member_name = out_member.upper()
    output_metadata.row_count = len(sorted_df)
    output_metadata.sort_keys = [str(name) for name in resolved_by]
    out_ds = Dataset(name=out_member, data=sorted_df, metadata=output_metadata)
    session.put_dataset(out_libref, out_member, out_ds)

    return StepResult(
        success=True,
        dataset_name=f"{out_libref}.{out_member}",
        rows_affected=len(sorted_df),
        notes=[f"Dataset {out_libref}.{out_member} created with {len(sorted_df)} observations."],
    )


def handle_proc_contents(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC CONTENTS — display dataset metadata."""
    data_name = proc.options.get("DATA", "")
    if not data_name:
        return StepResult(success=False, error="PROC CONTENTS requires DATA=")

    try:
        if "." in data_name:
            parts = data_name.split(".", 1)
            ds = session.get_dataset(parts[0], parts[1])
        else:
            ds = session.get_dataset("WORK", data_name)
    except KeyError:
        return StepResult(success=False, error=f"Dataset {data_name} not found")

    buf = io.StringIO()
    buf.write(f"\n{'=' * 60}\n")
    buf.write(f"  PROC CONTENTS: {ds.metadata.libref}.{ds.metadata.member_name}\n")
    buf.write(f"{'=' * 60}\n\n")
    buf.write(f"  Engine:     {ds.metadata.engine}\n")
    buf.write(f"  Observations: {ds.nrow}\n")
    buf.write(f"  Variables:    {ds.ncol}\n\n")

    buf.write(f"  {'#':<4} {'Variable':<20} {'Type':<12} {'Len':<6}\n")
    buf.write(f"  {'-' * 45}\n")
    for i, (logical_name, var) in enumerate(ds.metadata.variables.items(), 1):
        buf.write(f"  {i:<4} {var.name:<20} {var.dtype:<12} {var.length or 8:<6}\n")
    buf.write("\n")

    output = buf.getvalue()
    reporter.log(output)

    return StepResult(
        success=True,
        dataset_name=data_name,
        rows_affected=ds.nrow,
        output_messages=[output],
    )


def handle_proc_means(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC MEANS / SUMMARY — descriptive statistics."""
    data_name = proc.options.get("DATA", "")
    out_name = proc.options.get("OUT", "")
    noprint = proc.options.get("NOPRINT", False)
    maxdec = proc.options.get("MAXDEC", None)
    stat_names = [
        key for key in proc.options
        if key in ("N", "MEAN", "SUM", "MIN", "MAX", "STD", "MEDIAN", "Q1", "Q3")
    ]

    if not data_name:
        return StepResult(success=False, error="PROC MEANS requires DATA=")

    try:
        if "." in data_name:
            parts = data_name.split(".", 1)
            ds = session.get_dataset(parts[0], parts[1])
        else:
            ds = session.get_dataset("WORK", data_name)
    except KeyError:
        return StepResult(success=False, error=f"Dataset {data_name} not found")

    # Get VAR, CLASS, BY statements
    var_cols = []
    class_cols = []
    by_cols = []
    output_stmt = None
    for stmt in proc.statements:
        if isinstance(stmt, ClassNode):
            class_cols = [v.upper() for v in stmt.variables]
        elif isinstance(stmt, VarListNode):
            if not var_cols:
                var_cols = [v.upper() for v in stmt.variables]
        elif isinstance(stmt, ByNode):
            by_cols = [v.upper() for v in stmt.variables]
        elif isinstance(stmt, dict) and stmt.get("action") == "means_output":
            output_stmt = stmt
    group_cols = class_cols or by_cols

    # If no VAR specified, use all numeric columns
    if not var_cols:
        var_cols = [c for c in ds.data.columns if pd.api.types.is_numeric_dtype(ds.data[c])]

    # Case-insensitive column matching
    col_map = {c.upper(): c for c in ds.data.columns}
    valid_cols = [col_map[c] for c in var_cols if c in col_map]
    if not valid_cols:
        return StepResult(success=False, error="No valid numeric variables found")

    # Which stats to compute
    q1 = lambda series: series.quantile(0.25)
    q3 = lambda series: series.quantile(0.75)
    agg_map = {"N": "count", "MEAN": "mean", "SUM": "sum", "MIN": "min",
               "MAX": "max", "STD": "std", "MEDIAN": "median",
               "Q1": q1, "P25": q1, "Q3": q3, "P75": q3}
    if stat_names:
        aggs = {c: [agg_map[s] for s in stat_names if s in agg_map] for c in valid_cols}
    else:
        aggs = {c: ["count", "mean", "std", "min", "25%", "50%", "75%", "max"] for c in valid_cols}

    buf = io.StringIO()
    buf.write(f"\n{'=' * 60}\n")
    buf.write(f"  PROC MEANS: {ds.metadata.libref}.{ds.metadata.member_name}\n")
    buf.write(f"{'=' * 60}\n\n")

    if group_cols:
        # Grouped statistics
        valid_group = [col_map[c] for c in group_cols if c in col_map]
        if valid_group:
            grouped = ds.data.groupby(valid_group)
            if stat_names:
                # Named aggregations
                agg_dict = {}
                for vc in valid_cols:
                    for sn in stat_names:
                        fn = agg_map.get(sn)
                        if fn:
                            agg_dict[f"{vc}_{sn}"] = (vc, fn)
                stats_df = grouped.agg(**agg_dict)
            else:
                stats_df = grouped[valid_cols].describe()
            buf.write(stats_df.to_string())
            buf.write("\n")
        else:
            stats_df = ds.data[valid_cols].describe()
            buf.write(stats_df.to_string())
            buf.write("\n")
    else:
        if stat_names:
            parts = []
            for vc in valid_cols:
                col_stats = {}
                for sn in stat_names:
                    fn = agg_map.get(sn)
                    if fn == "count":
                        col_stats[sn] = ds.data[vc].count()
                    elif fn == "mean":
                        col_stats[sn] = ds.data[vc].mean()
                    elif fn == "sum":
                        col_stats[sn] = ds.data[vc].sum()
                    elif fn == "min":
                        col_stats[sn] = ds.data[vc].min()
                    elif fn == "max":
                        col_stats[sn] = ds.data[vc].max()
                    elif fn == "std":
                        col_stats[sn] = ds.data[vc].std()
                    elif fn == "median":
                        col_stats[sn] = ds.data[vc].median()
                    elif callable(fn):
                        col_stats[sn] = fn(ds.data[vc])
                parts.append(pd.Series(col_stats, name=vc))
            stats_df = pd.DataFrame(parts)
        else:
            stats_df = ds.data[valid_cols].describe()
        if maxdec is not None:
            stats_df = stats_df.round(int(maxdec))
        buf.write(stats_df.to_string())
        buf.write("\n")

    output = buf.getvalue()
    if not noprint:
        reporter.log(output)

    # OUTPUT OUT=ds STAT=newname ... statement
    if output_stmt and output_stmt.get("out"):
        stats_kv = output_stmt.get("stats", {})
        if not stats_kv:
            stats_kv = {"MEAN": "MEAN"}
        rows = []
        if group_cols:
            valid_group = [col_map[c] for c in group_cols if c in col_map]
            for key, block in ds.data.groupby(valid_group, dropna=False):
                if not isinstance(key, tuple):
                    key = (key,)
                row = dict(zip(valid_group, key))
                for stat_name, new_name in stats_kv.items():
                    fn = agg_map.get(stat_name)
                    if fn:
                        for vc in valid_cols:
                            row[new_name] = fn(block[vc]) if callable(fn) else getattr(block[vc], fn)()
                rows.append(row)
        else:
            row = {}
            for stat_name, new_name in stats_kv.items():
                fn = agg_map.get(stat_name)
                if fn:
                    for vc in valid_cols:
                        row[new_name] = fn(ds.data[vc]) if callable(fn) else getattr(ds.data[vc], fn)()
            rows.append(row)
        out_stmt_name = output_stmt["out"]
        if "." in out_stmt_name:
            parts = out_stmt_name.split(".", 1)
            o_libref, o_member = parts[0], parts[1]
        else:
            o_libref, o_member = "WORK", out_stmt_name
        o_ds = Dataset.from_dataframe(pd.DataFrame(rows), name=o_member, libref=o_libref)
        session.put_dataset(o_libref, o_member, o_ds)

    if out_name:
        if group_cols:
            valid_group = [col_map[c] for c in group_cols if c in col_map]
            if valid_group:
                grouped = ds.data.groupby(valid_group)
                out_data = grouped[valid_cols].agg(["count", "mean", "sum", "min", "max", "std"]).reset_index()
                # Flatten column names
                out_data.columns = [str(c) if not isinstance(c, tuple) else f"{c[0]}_{c[1]}" for c in out_data.columns]
            else:
                out_data = stats_df.reset_index()
        else:
            out_data = stats_df.reset_index()

        if "." in out_name:
            parts = out_name.split(".", 1)
            out_libref, out_name_only = parts[0], parts[1]
        else:
            out_libref, out_name_only = "WORK", out_name
        out_ds = Dataset.from_dataframe(out_data, name=out_name_only, libref=out_libref)
        session.put_dataset(out_libref, out_name_only, out_ds)

    return StepResult(
        success=True,
        rows_affected=ds.nrow,
        output_messages=[] if noprint else [output],
    )


def handle_proc_freq(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC FREQ — frequency tables with cross-tabulation support."""
    data_name = proc.options.get("DATA", "")
    if not data_name:
        return StepResult(success=False, error="PROC FREQ requires DATA=")

    try:
        if "." in data_name:
            parts = data_name.split(".", 1)
            ds = session.get_dataset(parts[0], parts[1])
        else:
            ds = session.get_dataset("WORK", data_name)
    except KeyError:
        return StepResult(success=False, error=f"Dataset {data_name} not found")

    buf = io.StringIO()
    buf.write(f"\n{'=' * 60}\n")
    buf.write(f"  PROC FREQ: {ds.metadata.libref}.{ds.metadata.member_name}\n")
    buf.write(f"{'=' * 60}\n")

    col_map = {c.upper(): c for c in ds.data.columns}

    # Collect FreqTableSpec from statements
    table_specs: list[FreqTableSpec] = []
    for stmt in proc.statements:
        if isinstance(stmt, list):
            table_specs.extend([s for s in stmt if isinstance(s, FreqTableSpec)])
        elif isinstance(stmt, FreqTableSpec):
            table_specs.append(stmt)

    if not table_specs:
        return StepResult(success=False, error="PROC FREQ requires TABLES statement")

    for spec in table_specs:
        # Resolve variable names (case-insensitive)
        vars_resolved = []
        for v in spec.var_names:
            actual = col_map.get(v.upper())
            if actual:
                vars_resolved.append(actual)
            else:
                return StepResult(success=False, error=f"Variable {v} not found in dataset")

        if not vars_resolved:
            continue

        opts = spec.options
        include_missing = opts.get("MISSING", False)
        show_row = not opts.get("NOROW", False)
        show_col = not opts.get("NOCOL", False)
        show_pct = not opts.get("NOPERCENT", False)

        if len(vars_resolved) == 1:
            _freq_one_way(buf, ds.data, vars_resolved[0], include_missing)
        else:
            _freq_crosstab(buf, ds.data, vars_resolved, include_missing,
                           show_row, show_col, show_pct)

    output = buf.getvalue()
    reporter.log(output)

    return StepResult(success=True, rows_affected=ds.nrow, output_messages=[output])


def _freq_one_way(buf: io.StringIO, df: pd.DataFrame, var_name: str,
                  include_missing: bool) -> None:
    """Write a one-way frequency table."""
    if include_missing:
        freq = df[var_name].value_counts(dropna=False).sort_index()
    else:
        freq = df[var_name].dropna().value_counts().sort_index()

    total = freq.sum()
    cum_freq = 0
    cum_pct = 0.0

    buf.write(f"\n\n  Variable: {var_name}\n")
    buf.write(f"  {'Value':<20} {'Frequency':>10} {'Percent':>10} {'Cum Freq':>10} {'Cum Pct':>10}\n")
    buf.write(f"  {'-' * 62}\n")

    for val, count in freq.items():
        pct = count / total * 100 if total > 0 else 0
        cum_freq += count
        cum_pct += pct
        val_str = _fmt_freq_val(val)
        buf.write(f"  {val_str:<20} {count:>10} {pct:>10.2f} {cum_freq:>10} {cum_pct:>10.2f}\n")

    buf.write(f"  {'Total':<20} {total:>10} {'100.00':>10}\n")


def _freq_crosstab(buf: io.StringIO, df: pd.DataFrame, vars_resolved: list[str],
                   include_missing: bool, show_row: bool, show_col: bool,
                   show_pct: bool) -> None:
    """Write a cross-tabulation frequency table."""
    row_var = vars_resolved[0]
    col_var = vars_resolved[1]

    if include_missing:
        ct = pd.crosstab(df[row_var], df[col_var], margins=True, margins_name="Total",
                         dropna=False)
    else:
        ct = pd.crosstab(df[row_var], df[col_var], margins=True, margins_name="Total")

    buf.write(f"\n\n  Table {row_var} * {col_var}\n")

    # Format the crosstab with frequency and optional percents
    n_rows = len(ct.index) - 1  # exclude Total row
    n_cols = len(ct.columns) - 1  # exclude Total column

    # Print frequency table
    buf.write(f"\n  Frequency  |{' ' * 5}\n")
    buf.write(f"  {'':20} |")
    for col in ct.columns:
        buf.write(f" {_fmt_freq_val(col):>10}")
    buf.write("\n")
    buf.write(f"  {'-' * 20}-+-{'-' * (11 * len(ct.columns))}\n")

    grand_total = ct.loc["Total", "Total"] if "Total" in ct.index and "Total" in ct.columns else ct.values.sum()

    for idx in ct.index:
        val_str = _fmt_freq_val(idx)
        buf.write(f"  {val_str:<20} |")
        for col in ct.columns:
            count = ct.loc[idx, col]
            buf.write(f" {int(count):>10}")
        buf.write("\n")

    # Percent table
    if show_pct and grand_total > 0:
        buf.write(f"\n  Percent    |{' ' * 5}\n")
        buf.write(f"  {'':20} |")
        for col in ct.columns:
            buf.write(f" {_fmt_freq_val(col):>10}")
        buf.write("\n")
        buf.write(f"  {'-' * 20}-+-{'-' * (11 * len(ct.columns))}\n")

        for idx in ct.index:
            val_str = _fmt_freq_val(idx)
            buf.write(f"  {val_str:<20} |")
            for col in ct.columns:
                pct = ct.loc[idx, col] / grand_total * 100
                buf.write(f" {pct:>10.2f}")
            buf.write("\n")

    # Row percent table
    if show_row and n_rows > 0:
        buf.write(f"\n  Row Pct    |{' ' * 5}\n")
        buf.write(f"  {'':20} |")
        for col in ct.columns:
            buf.write(f" {_fmt_freq_val(col):>10}")
        buf.write("\n")
        buf.write(f"  {'-' * 20}-+-{'-' * (11 * len(ct.columns))}\n")

        for idx in ct.index:
            val_str = _fmt_freq_val(idx)
            buf.write(f"  {val_str:<20} |")
            row_total = ct.loc[idx, "Total"] if "Total" in ct.columns else ct.loc[idx].sum()
            for col in ct.columns:
                if row_total > 0:
                    pct = ct.loc[idx, col] / row_total * 100
                else:
                    pct = 0.0
                buf.write(f" {pct:>10.2f}")
            buf.write("\n")

    # Col percent table
    if show_col and n_cols > 0:
        buf.write(f"\n  Col Pct    |{' ' * 5}\n")
        buf.write(f"  {'':20} |")
        for col in ct.columns:
            buf.write(f" {_fmt_freq_val(col):>10}")
        buf.write("\n")
        buf.write(f"  {'-' * 20}-+-{'-' * (11 * len(ct.columns))}\n")

        for idx in ct.index:
            val_str = _fmt_freq_val(idx)
            buf.write(f"  {val_str:<20} |")
            for col in ct.columns:
                col_total = ct.loc["Total", col] if "Total" in ct.index else ct[col].sum()
                if col_total > 0:
                    pct = ct.loc[idx, col] / col_total * 100
                else:
                    pct = 0.0
                buf.write(f" {pct:>10.2f}")
            buf.write("\n")


def handle_proc_append(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC APPEND — append rows from one dataset to another."""
    base_name = proc.options.get("BASE", "")
    data_name = proc.options.get("DATA", "")

    if not base_name:
        return StepResult(success=False, error="PROC APPEND requires BASE=")
    if not data_name:
        return StepResult(success=False, error="PROC APPEND requires DATA=")

    try:
        if "." in base_name:
            parts = base_name.split(".", 1)
            base_ds = session.get_dataset(parts[0], parts[1])
        else:
            base_ds = session.get_dataset("WORK", base_name)
    except KeyError:
        return StepResult(success=False, error=f"Dataset {base_name} not found")

    try:
        if "." in data_name:
            parts = data_name.split(".", 1)
            data_ds = session.get_dataset(parts[0], parts[1])
        else:
            data_ds = session.get_dataset("WORK", data_name)
    except KeyError:
        return StepResult(success=False, error=f"Dataset {data_name} not found")

    # Align columns — use base columns
    base_cols = list(base_ds.data.columns)
    data_df = data_ds.data.copy()
    for col in base_cols:
        if col not in data_df.columns:
            data_df[col] = None
    data_df = data_df[[c for c in base_cols if c in data_df.columns]]

    new_df = pd.concat([base_ds.data, data_df], ignore_index=True)
    out_ds = Dataset.from_dataframe(new_df, name=base_ds.metadata.member_name,
                                     libref=base_ds.metadata.libref)

    if "." in base_name:
        parts = base_name.split(".", 1)
        session.put_dataset(parts[0], parts[1], out_ds)
    else:
        session.put_dataset("WORK", base_name, out_ds)

    return StepResult(
        success=True,
        rows_affected=len(data_ds.data),
        notes=[f"{len(data_ds.data)} observations appended to {base_name}."],
    )


def handle_proc_datasets(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC DATASETS — manage datasets (DELETE, MODIFY, CONTENTS)."""
    libref = proc.options.get("LIBRARY", "WORK")
    if "." in libref:
        libref = libref.split(".", 1)[0]

    notes = []
    if proc.options.get("KILL"):
        backend = session.storage.get_backend(libref)
        if backend is None:
            return StepResult(success=False, error=f"Library {libref} is not defined")
        deleted = list(backend.list_datasets())
        for name in deleted:
            backend.delete(name)
        if deleted:
            notes.append(f"Deleted all {len(deleted)} members from {libref}.")

    for stmt in proc.statements:
        if isinstance(stmt, list):
            for s in stmt:
                result = _exec_ds_stmt(s, libref, session, reporter)
                if not result.success:
                    return result
                notes.extend(result.notes)
        else:
            result = _exec_ds_stmt(stmt, libref, session, reporter)
            if not result.success:
                return result
            notes.extend(result.notes)

    return StepResult(success=True, notes=notes)


def _exec_ds_stmt(stmt: Any, libref: str, session: Session, reporter: Reporter) -> StepResult:
    """Execute a single PROC DATASETS sub-statement."""
    if not isinstance(stmt, dict):
        return StepResult(success=True)

    action = stmt.get("action", "")

    if action == "delete":
        names = stmt.get("names", [])
        deleted = []
        for name in names:
            name_upper = name.upper() if isinstance(name, str) else str(name).upper()
            try:
                if "." in name_upper:
                    parts = name_upper.split(".", 1)
                    backend, ds_name = session.storage.resolve(parts[0], parts[1])
                else:
                    backend, ds_name = session.storage.resolve(libref, name_upper)
                if backend.exists(ds_name):
                    backend.delete(ds_name)
                    deleted.append(name_upper)
            except (KeyError, Exception):
                pass
        if deleted:
            return StepResult(success=True, notes=[f"Deleted: {', '.join(deleted)}"])
        return StepResult(success=True)

    if action == "contents":
        data_name = stmt.get("name", "")
        if data_name:
            try:
                if "." in data_name:
                    parts = data_name.split(".", 1)
                    session.get_dataset(parts[0], parts[1])
                else:
                    session.get_dataset(libref, data_name)
                return handle_proc_contents(
                    ProcNode(proc_name="CONTENTS", options={"DATA": data_name}),
                    session, reporter,
                )
            except KeyError:
                return StepResult(success=False, error=f"Dataset {data_name} not found")
        return StepResult(success=True)

    if action == "modify":
        # MODIFY with optional RENAME — simplified implementation
        return StepResult(success=True)

    return StepResult(success=True)


def _fmt_freq_val(val: Any) -> str:
    """Format a value for frequency table display."""
    import math
    if val is None:
        return "."
    if isinstance(val, float) and math.isnan(val):
        return "."
    return str(val)


def handle_proc_import(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC IMPORT — import data from file."""
    datafile = proc.options.get("DATAFILE", "")
    out_name = proc.options.get("OUT", "")
    dbms = proc.options.get("DBMS", "CSV").lower()
    delimiter = proc.options.get("DELIMITER", ",")
    getnames = proc.options.get("GETNAMES", "YES").upper() == "YES"

    if not datafile:
        return StepResult(success=False, error="PROC IMPORT requires DATAFILE=")
    if not out_name:
        return StepResult(success=False, error="PROC IMPORT requires OUT=")

    try:
        import os
        filepath = datafile
        if not os.path.isabs(filepath):
            filepath = os.path.abspath(filepath)

        if dbms == "csv" or dbms == "dlm" or dbms == "tab":
            sep = delimiter if dbms != "tab" else "\t"
            header = 0 if getnames else None
            df = pd.read_csv(filepath, sep=sep, header=header)
        elif dbms == "xlsx" or dbms == "excel":
            df = pd.read_excel(filepath, header=0 if getnames else None)
        else:
            return StepResult(success=False, error=f"Unsupported DBMS type: {dbms}")

        # Parse output dataset name
        if "." in out_name:
            parts = out_name.split(".", 1)
            out_libref = parts[0].upper()
            out_member = parts[1].upper()
        else:
            out_libref = "WORK"
            out_member = out_name.upper()

        out_ds = Dataset.from_dataframe(df, name=out_member, libref=out_libref)
        session.put_dataset(out_libref, out_member, out_ds)

        return StepResult(
            success=True,
            dataset_name=f"{out_libref}.{out_member}",
            rows_affected=len(df),
            notes=[f"Dataset {out_libref}.{out_member} created with {len(df)} observations and {len(df.columns)} variables."],
        )
    except FileNotFoundError:
        return StepResult(success=False, error=f"File not found: {datafile}")
    except Exception as e:
        return StepResult(success=False, error=f"PROC IMPORT error: {e}")


def handle_proc_export(proc: ProcNode, session: Session, reporter: Reporter) -> StepResult:
    """PROC EXPORT — export data to file."""
    data_ref = proc.options.get("DATA", "")
    outfile = proc.options.get("OUTFILE", "")
    dbms = proc.options.get("DBMS", "CSV").lower()
    delimiter = proc.options.get("DELIMITER", ",")
    replace = bool(proc.options.get("REPLACE", False))
    use_labels = bool(proc.options.get("LABEL", False))
    sheet_name = str(proc.options.get("SHEET", "Sheet1"))

    if not data_ref:
        return StepResult(success=False, error="PROC EXPORT requires DATA=")
    if not outfile:
        return StepResult(success=False, error="PROC EXPORT requires OUTFILE=")

    try:
        if isinstance(data_ref, DatasetRefNode):
            data_libref = data_ref.libref
            data_member = data_ref.name
            data_options = data_ref.options
        else:
            data_name = str(data_ref)
            if "." in data_name:
                data_libref, data_member = data_name.split(".", 1)
            else:
                data_libref, data_member = "WORK", data_name
            data_options = []
        display_name = f"{data_libref}.{data_member}"
        ds = session.get_dataset(data_libref, data_member)
        ds = _apply_export_dataset_options(ds, data_options, session)

        import os
        filepath = outfile
        if not os.path.isabs(filepath):
            filepath = os.path.abspath(filepath)

        if os.path.exists(filepath) and not replace:
            return StepResult(
                success=False,
                error=(
                    f"Output file already exists: {filepath}. "
                    "Specify REPLACE to overwrite it."
                ),
            )

        export_df = ds.data.copy()
        if use_labels:
            label_mapping = {}
            for column in export_df.columns:
                metadata = ds.metadata.get_variable(str(column))
                if metadata is not None and metadata.label:
                    label_mapping[column] = metadata.label
            if label_mapping:
                export_df = export_df.rename(columns=label_mapping)

        if dbms == "csv" or dbms == "dlm":
            csv_dataframe(ds).to_csv(filepath, sep=delimiter, index=False)
        elif dbms == "tab":
            csv_dataframe(ds).to_csv(filepath, sep="\t", index=False)
        elif dbms == "xlsx" or dbms == "excel":
            try:
                export_df.to_excel(filepath, index=False, sheet_name=sheet_name)
                from openpyxl import load_workbook
                from openpyxl.utils import get_column_letter

                workbook = load_workbook(filepath)
                worksheet = workbook[sheet_name]
                for index, column in enumerate(export_df.columns, 1):
                    values = [str(column), *export_df.iloc[:, index - 1].fillna("").astype(str)]
                    width = min(max(len(value) for value in values) + 2, 60)
                    worksheet.column_dimensions[get_column_letter(index)].width = width
                workbook.save(filepath)
            except ImportError:
                return StepResult(
                    success=False,
                    error=(
                        f"Cannot export {filepath}: Excel export requires the "
                        "optional dependency openpyxl. "
                        "Install SASLite with the excel extra: pip install "
                        "'saslite[excel]'"
                    ),
                )
        else:
            return StepResult(success=False, error=f"Unsupported DBMS type: {dbms}")

        return StepResult(
            success=True,
            rows_affected=ds.nrow,
            notes=[f"Dataset {display_name} exported to {filepath} ({ds.nrow} observations)."],
        )
    except KeyError:
        return StepResult(success=False, error=f"Dataset {data_ref} not found")
    except Exception as e:
        return StepResult(success=False, error=f"PROC EXPORT error: {e}")


def _apply_export_dataset_options(
    ds: Dataset,
    options: list[Any],
    session: Session,
) -> Dataset:
    """Apply input data set options before PROC EXPORT writes the file."""
    result = ds.copy()
    registry = build_default_registry()

    for option in options:
        if not isinstance(option, dict):
            continue
        if "KEEP" in option:
            keep = {str(name).upper() for name in option["KEEP"]}
            result = result.select_columns([
                column for column in result.data.columns
                if str(column).upper() in keep
            ])
        if "DROP" in option:
            drop = {str(name).upper() for name in option["DROP"]}
            result = result.select_columns([
                column for column in result.data.columns
                if str(column).upper() not in drop
            ])
        if "RENAME" in option:
            columns = {str(column).upper(): column for column in result.data.columns}
            rename = {
                columns[str(old).upper()]: new
                for old, new in option["RENAME"].items()
                if str(old).upper() in columns
            }
            result = result.rename_columns(rename)
        if "WHERE" in option:
            condition = option["WHERE"]
            columns = {str(column).upper(): column for column in result.data.columns}
            mask: list[bool] = []
            for _, row in result.data.iterrows():
                evaluator = ExpressionEvaluator(
                    var_getter=lambda name, current=row: current.get(
                        columns.get(name.upper(), name)
                    ),
                    session=session,
                )
                for function_name in registry.names:
                    function = registry.get(function_name)
                    if function is not None:
                        evaluator.register_function(function_name, function)
                mask.append(sas_bool(evaluator.evaluate(condition)))
            result.data = result.data.loc[mask].reset_index(drop=True)

    return result
